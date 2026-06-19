"""
Admin routes: dashboard de monitoreo y exportacion CSV.
Acceso: /resultados
"""
import csv
import io
import os
import socket
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Request, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, FileResponse
from fastapi.templating import Jinja2Templates

from db import get_conn, create_session, save_answer, mark_complete
from questions import QUESTIONS, AREA_LABELS, compute_result

PORT = int(os.environ.get("PORT", 8777))


def get_local_ip() -> str:
    """Detecta la mejor IP de red disponible (Walmart LAN o Hotspot movil)."""
    candidates = []
    try:
        import subprocess
        # ipconfig es el metodo mas confiable en Windows
        out = subprocess.check_output("ipconfig", text=True, timeout=3)
        for line in out.splitlines():
            if "IPv4" in line:
                ip = line.split(":")[-1].strip()
                if ip and not ip.startswith("127.") and not ip.startswith("169.254"):
                    candidates.append(ip)
    except Exception:
        pass

    if not candidates:
        # Fallback: socket trick
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("10.0.0.0", 80))
            candidates.append(s.getsockname()[0])
            s.close()
        except Exception:
            pass

    # Preferir IPs de hotspot (192.168.x.x) sobre IPs corporativas (10.x.x.x)
    for ip in candidates:
        if ip.startswith("192.168."):
            return ip
    return candidates[0] if candidates else "localhost"


def get_survey_url() -> tuple[str, str]:
    """Retorna (url, tipo_red). Respeta PUBLIC_URL si esta en la nube."""
    public = os.environ.get("PUBLIC_URL", "").rstrip("/")
    if public:
        return public + "/", "cloud"
    ip = get_local_ip()
    url = f"http://{ip}:{PORT}/"
    if ip == "localhost":
        tipo = "local"
    elif ip.startswith("192.168."):
        tipo = "hotspot"
    else:
        tipo = "lan"
    return url, tipo

TEMPL     = Path(__file__).parent / "templates"
router    = APIRouter()
templates = Jinja2Templates(directory=TEMPL)

Q_MAP = {str(q["id"]): q for q in QUESTIONS}


def _all_sessions() -> list[dict]:
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT id, rut, razon_social, email, created_at, completed "
            "FROM sessions ORDER BY created_at DESC"
        )
        sessions = cur.fetchall()
        cur.execute("SELECT session_id, question_id, answer FROM answers")
        all_answers = cur.fetchall()

    ans_map: dict[str, dict] = {}
    for a in all_answers:
        ans_map.setdefault(a["session_id"], {})[str(a["question_id"])] = a["answer"]

    rows = []
    for s in sessions:
        sid     = s["id"]
        answers = ans_map.get(sid, {})
        result  = compute_result(answers)
        rows.append({
            "id":           sid,
            "rut":          s["rut"]          or "",
            "razon_social": s["razon_social"] or "",
            "email":        s["email"]        or "",
            "created_at":   str(s["created_at"]) if s["created_at"] else "",
            "completed":    bool(s["completed"]),
            "answers":      answers,
            **result,
        })
    return rows


def _stats(rows: list[dict]) -> dict:
    total       = len(rows)
    completadas = sum(1 for r in rows if r["completed"])
    intermedias = sum(1 for r in rows if r["completed"] and r["segment"] == "intermedia")
    maduras     = sum(1 for r in rows if r["completed"] and r["segment"] == "madura")
    pct_inter   = round(intermedias / completadas * 100) if completadas else 0
    pct_mad     = round(maduras    / completadas * 100) if completadas else 0
    avg_score   = round(sum(r["pct"] for r in rows if r["completed"]) / completadas) if completadas else 0
    return {
        "total": total, "completadas": completadas,
        "intermedias": intermedias, "maduras": maduras,
        "pct_inter": pct_inter, "pct_mad": pct_mad,
        "avg_score": avg_score,
        "en_progreso": total - completadas,
    }


# ── Reset datos (para limpiar pruebas) ────────────────────────────────────
@router.post("/resultados/reset")
async def reset_data():
    """Borra TODOS los datos. Solo para uso interno/testing."""
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM answers")
        cur.execute("DELETE FROM sessions")
    return RedirectResponse(url="/resultados", status_code=303)


# ── Descarga de Excel Manual (Plan B) ──────────────────────────────────────────
@router.get("/descargar-excel-manual")
async def descargar_excel_manual():
    """Permite a los proveedores descargar la plantilla de Excel
    directamente haciendo clic en el PDF o ingresando a la URL."""
    try:
        file_path = Path(__file__).parent / "ENCUESTA_MANUAL_PYME.xlsx"
        return FileResponse(
            path=file_path,
            filename="ENCUESTA_MANUAL_PYME.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        print(f"Error sirviendo Excel de descarga: {e}")
        return HTMLResponse("Archivo no disponible en este momento", status_code=404)


# ── Descarga de Guía PDF (Acceso Rápido) ─────────────────────────────────────
@router.get("/descargar-pdf-guia")
async def descargar_pdf_guia():
    """Permite descargar la Guía PDF oficial del Proveedor en el Dashboard."""
    try:
        file_path = Path(__file__).parent.parent / "GUIA_ENCUESTA_PYME.pdf"
        return FileResponse(
            path=file_path,
            filename="GUIA_ENCUESTA_PYME.pdf",
            media_type="application/pdf"
        )
    except Exception as e:
        print(f"Error sirviendo Guia PDF de descarga: {e}")
        return HTMLResponse("Archivo no disponible en este momento", status_code=404)


# ── Eliminación de PyME Manual (Tachito de Basura) ───────────────────────────
@router.post("/resultados/eliminar/{session_id}")
async def eliminar_sesion(session_id: str):
    """Elimina permanentemente una sesión y todas sus respuestas asociadas,
    tanto en SQLite (local) como en PostgreSQL (producción)."""
    try:
        from db import get_conn, _ph
        ph = _ph(1)
        with get_conn() as conn:
            cur = conn.cursor()
            # Eliminar respuestas en cascada primero
            cur.execute(f"DELETE FROM answers WHERE session_id = {ph}", (session_id,))
            # Eliminar la sesión
            cur.execute(f"DELETE FROM sessions WHERE id = {ph}", (session_id,))
        return RedirectResponse(url="/resultados?success=eliminado", status_code=303)
    except Exception as e:
        print(f"Error eliminando sesión del panel: {e}")
        return RedirectResponse(url="/resultados?error=eliminar_error", status_code=303)


# ── Importación de Excel Manual (Plan B) ───────────────────────────────────────
@router.post("/resultados/importar-excel")
async def importar_excel_manual(file: UploadFile = File(...)):
    """Recibe un archivo Excel del Plan B (ENCUESTA_MANUAL_PYME.xlsx),
    lo procesa, lo guarda en la base de datos y calcula sus métricas."""
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active

        # Lectura de identificación de la empresa (C5, C6, C7)
        rut = str(ws["C5"].value or "").strip()
        razon_social = str(ws["C6"].value or "").strip()
        email = str(ws["C7"].value or "").strip()

        if not rut or not razon_social:
            return RedirectResponse(url="/resultados?error=excel_invalido", status_code=303)

        # Crear sesión en la base de datos
        sid = create_session(rut, razon_social, email)

        # Recorremos filas de la 10 a la 34 (las 25 preguntas)
        for row in range(10, 35):
            qid = ws.cell(row=row, column=1).value
            respuesta_user = str(ws.cell(row=row, column=3).value or "").strip()

            if not qid:
                continue

            mapped_ans = respuesta_user
            q = Q_MAP.get(str(qid))
            if q and q["options"]:
                # Mapeo por coincidencia inteligente de texto
                matched = False
                for opt in q["options"]:
                    if respuesta_user.lower() in opt["label"].lower() or opt["label"].lower() in respuesta_user.lower():
                        mapped_ans = opt["value"]
                        matched = True
                        break
                if not matched and respuesta_user:
                    # Si escribió algo pero no hizo match perfecto, dejamos el primer valor por defecto
                    mapped_ans = q["options"][0]["value"]
            
            save_answer(sid, int(qid), mapped_ans)

        # Marcamos la sesión como completada para que se compute en el dashboard
        mark_complete(sid)
        return RedirectResponse(url="/resultados?success=excel_importado", status_code=303)

    except Exception as e:
        print(f"Error importando Excel manual: {e}")
        return RedirectResponse(url="/resultados?error=excel_error", status_code=303)


# ── Dashboard ───────────────────────────────────────────────────────────────
@router.get("/resultados", response_class=HTMLResponse)
async def dashboard(request: Request, q: str = ""):
    all_rows = _all_sessions()
    rows     = all_rows
    if q:
        q_low = q.lower()
        rows  = [r for r in rows
                 if q_low in r["rut"].lower()
                 or q_low in r["razon_social"].lower()
                 or q_low in r["email"].lower()]
    stats = _stats(all_rows)
    survey_url, net_type = get_survey_url()
    
    from questions import QUESTIONS, BLOCKS
    return templates.TemplateResponse(request, "admin.html", {
        "rows":        rows,
        "stats":       stats,
        "q":           q,
        "area_labels": AREA_LABELS,
        "survey_url":  survey_url,
        "net_type":    net_type,
        "now":         datetime.now().strftime("%H:%M:%S"),
        "questions":   QUESTIONS,
        "blocks":      BLOCKS,
    })


# ── Partial: stats + tabla (para HTMX polling) ───────────────────────────────
@router.get("/resultados/live", response_class=HTMLResponse)
async def live_partial(request: Request, q: str = ""):
    all_rows = _all_sessions()
    rows     = all_rows
    if q:
        q_low = q.lower()
        rows  = [r for r in rows
                 if q_low in r["rut"].lower()
                 or q_low in r["razon_social"].lower()
                 or q_low in r["email"].lower()]
    stats = _stats(all_rows)
    return templates.TemplateResponse(request, "admin_live.html", {
        "rows":        rows,
        "stats":       stats,
        "q":           q,
        "area_labels": AREA_LABELS,
        "now":         datetime.now().strftime("%H:%M:%S"),
    })


# ── Detalle por sesion ────────────────────────────────────────────────────────
@router.get("/resultados/{session_id}", response_class=HTMLResponse)
async def detail(request: Request, session_id: str):
    rows = _all_sessions()
    row  = next((r for r in rows if r["id"] == session_id), None)
    if not row:
        return HTMLResponse("Sesion no encontrada", status_code=404)
    return templates.TemplateResponse(request, "admin_detail.html", {
        "s":           row,
        "questions":   QUESTIONS,
        "area_labels": AREA_LABELS,
    })


# ── Export CSV ────────────────────────────────────────────────────────────────
@router.get("/resultados/export/csv")
async def export_csv():
    rows = _all_sessions()

    output = io.StringIO()
    writer = csv.writer(output)

    # Encabezados
    headers = [
        "RUT", "Razon Social", "Email", "Fecha", "Completada",
        "Segmento", "Score %", "Puntaje", "Max Puntaje", "Areas Debiles",
    ]
    for q in QUESTIONS:
        headers.append(f"P{q['id']} - {q['text'][:40]}")
    writer.writerow(headers)

    # Filas
    for r in rows:
        row_data = [
            r["rut"],
            r["razon_social"],
            r["email"],
            r["created_at"][:16] if r["created_at"] else "",
            "Si" if r["completed"] else "No",
            r["segment_label"],
            r["pct"],
            r["score"],
            r["max_score"],
            " | ".join(AREA_LABELS[a]["label"] for a in r["weak_areas"] if a in AREA_LABELS),
        ]
        for q in QUESTIONS:
            row_data.append(r["answers"].get(str(q["id"]), ""))
        writer.writerow(row_data)

    output.seek(0)
    fecha = datetime.now().strftime("%Y%m%d_%H%M")
    return StreamingResponse(
        iter([output.getvalue().encode("utf-8-sig")]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=pymes_cuestionario_{fecha}.csv"},
    )


# ── Helpers Excel ────────────────────────────────────────────────────────────
def _header_fill():  return PatternFill("solid", fgColor="002060")
def _spark_fill():   return PatternFill("solid", fgColor="FFC000")
def _alt_fill():     return PatternFill("solid", fgColor="EEF3FD")
def _green_fill():   return PatternFill("solid", fgColor="D4EDDA")
def _yellow_fill():  return PatternFill("solid", fgColor="FFF3CD")
def _thin_border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_font():  return Font(name="Calibri", bold=True, color="FFFFFF", size=11)
def _body_font():    return Font(name="Calibri", size=10)
def _bold_font():    return Font(name="Calibri", bold=True, size=10, color="002060")

def _set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def _style_header_row(ws, row: int, n_cols: int):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = _header_fill()
        cell.font      = _header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _thin_border()


# ── Export Excel completo (formateado) ──────────────────────────────────────────
@router.get("/resultados/export/excel")
async def export_excel():
    rows  = _all_sessions()
    fecha = datetime.now().strftime("%Y%m%d_%H%M")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Respuestas PyME"
    ws.freeze_panes = "A2"  # congela header

    # — Encabezados —
    base_headers = [
        "RUT", "Razón Social", "Email", "Fecha",
        "Completada", "Segmento", "Score %",
    ]
    q_headers = [f"P{q['id']} - {q['text'][:35]}" for q in QUESTIONS]
    all_headers = base_headers + q_headers

    ws.append(all_headers)
    _style_header_row(ws, 1, len(all_headers))
    ws.row_dimensions[1].height = 32

    # — Filas de datos —
    seg_fills = {"madura": _green_fill(), "intermedia": _yellow_fill()}

    for i, r in enumerate(rows, start=2):
        row_data = [
            r["rut"],
            r["razon_social"] or "Sin nombre",
            r["email"],
            r["created_at"][:16] if r["created_at"] else "",
            "Sí" if r["completed"] else "No",
            r["segment_label"] if r["completed"] else "En progreso",
            r["pct"]           if r["completed"] else "",
        ] + [r["answers"].get(str(q["id"]), "") for q in QUESTIONS]

        ws.append(row_data)

        # Fondo alternado + color segmento
        row_fill = seg_fills.get(r.get("segment", ""), _alt_fill() if i % 2 == 0 else None)
        for c in range(1, len(all_headers) + 1):
            cell = ws.cell(row=i, column=c)
            cell.font      = _bold_font() if c == 2 else _body_font()
            cell.border    = _thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if row_fill and c in (5, 6, 7):  # solo columnas de estado
                cell.fill = row_fill
            elif i % 2 == 0:
                cell.fill = _alt_fill()

    # — Anchos de columna —
    fixed = {"A": 16, "B": 28, "C": 28, "D": 16, "E": 11, "F": 16, "G": 10}
    _set_col_widths(ws, fixed)
    for c in range(8, len(all_headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 32

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=pymes_respuestas_{fecha}.xlsx"},
    )


# ── Export Excel contactos (RUT + Razón Social + Email) ─────────────────────
@router.get("/resultados/export/contactos")
async def export_contactos():
    rows  = _all_sessions()
    fecha = datetime.now().strftime("%Y%m%d_%H%M")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contactos PyME"
    ws.freeze_panes = "A2"

    # Titulo decorativo
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value     = "Contactos Programa PyME • Walmart Chile"
    title_cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    title_cell.fill      = _spark_fill()
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header
    headers = ["RUT", "Razón Social", "Email", "Fecha"]
    ws.append(headers)
    _style_header_row(ws, 2, len(headers))
    ws.row_dimensions[2].height = 24

    # Filas
    for i, r in enumerate(rows, start=3):
        ws.append([
            r["rut"],
            r["razon_social"] or "Sin nombre",
            r["email"],
            r["created_at"][:10] if r["created_at"] else "",
        ])
        for c in range(1, 5):
            cell = ws.cell(row=i, column=c)
            cell.font      = _bold_font() if c == 2 else _body_font()
            cell.border    = _thin_border()
            cell.alignment = Alignment(vertical="center")
            if i % 2 == 0:
                cell.fill = _alt_fill()

    _set_col_widths(ws, {"A": 18, "B": 32, "C": 34, "D": 14})

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=pymes_contactos_{fecha}.xlsx"},
    )
